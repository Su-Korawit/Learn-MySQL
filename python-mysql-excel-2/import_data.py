from openpyxl import load_workbook # Excel
import mysql.connector # Database
from unicodedata import category

# Excel

workbook = load_workbook('imported_02.xlsx')
sheet = workbook.active

# Database

db = mysql.connector.connect(
    host = 'localhost',
    port = 3306,
    user = 'root',
    password = 'root64495',
    database = 'goft_want_to_but'
)
cursor = db.cursor()

# โหลดข้อมูลประเภทสินค้าทั้งหมด
sql_select_categories = '''
    SELECT *
    FROM categories
'''

cursor.execute(sql_select_categories)
categories = cursor.fetchall()

# เปรียนแทบข้อมูบประเภทสินค้า อันไหนไม่มี ให้เพิ่มลงไป

categories_values = []
for row in sheet.iter_rows(min_row = 2, values_only = True):
    is_new = True
    category = row[3]

    for c in categories:
        if category == c[1]:
            is_new = False
            break

    if is_new:
        print((category, ))
        categories_values.append((category, ))

if len(categories_values) > 0:
    sql_insert_categories = '''
        INSERT INTO categories (title)
        VALUES (%s)
    '''

    cursor.executemany(sql_insert_categories, categories_values)
    db.commit()
    print('ADD', str(cursor.rowcount), 'ROW')
else:
    print('NO PRODUCT')

# โหลดข้อมูลประเภทสินค้าทั้งหมด อีกครั้ง

cursor.execute(sql_select_categories)
categories = cursor.fetchall()

# เชื่อมต่อ categoriy_id กับสินค้าใหม่ของเรา แล้วเพิ่มลงไป
product_values = []
for row in sheet.iter_rows(min_row = 2, values_only = True):
    category_title = row[3]
    category_id = 'NULL'

    for c in categories :
        if category_title == c[1]:
            category_id = c[0]
            break

    product = (row[0], row[1], row[2], category_id)
    print(product)
    product_values.append(product)

sql_insert_products = '''
    INSERT INTO products (title, price, is_necessary, category_id)
    VALUES (%s, %s, %s, %s)
'''

cursor.executemany(sql_insert_products, product_values)
db.commit()
print('ADD', str(cursor.rowcount), 'ROW')

cursor.close()
db.close()