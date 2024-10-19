from openpyxl import load_workbook # Excel
import mysql.connector # Database

# Excel

workbook = load_workbook('imported.xlsx')
sheet = workbook.active

values = []
for row in sheet.iter_rows(min_row = 2, values_only = True):
    print(row)
    values.append(row)

# Database

db = mysql.connector.connect(
    host = 'localhost',
    port = 3306,
    user = 'root',
    password = 'root64495',
    database = 'goft_want_to_but'
)

cursor = db.cursor()
sql = '''
    INSERT INTO products (title, price, is_necessary)
    VALUES (%s, %s, %s);
'''

cursor.executemany(sql, values)
db.commit()
print('ADD DATE COUNT', str(cursor.rowcount), 'ROW')